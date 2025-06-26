from openpyxl import load_workbook
import pyautogui
import time
import keyboard
import pygetwindow as gw

planilha = load_workbook("C:\\Users\\vitor.santos\\Documents\\Estudos\\Python\\Automação\\Html\\python\\produtos - Copia.xlsx", data_only=True)
pagina = planilha['Plan1']

produtolist = []
qtdlist = []
armazenlist = []
documento = 123456
for dados in pagina.iter_rows(values_only=True):
    produtolist.append(dados[0])
    qtdlist.append(dados[1])
    armazenlist.append(dados[2])

print(f"{produtolist} \n \n{qtdlist} \n \n{armazenlist}")

def clicar(vezes, tecla):
    i = 0
    while i < vezes: 
        time.sleep(0.01)
        keyboard.send(tecla)
        i = i + 1  

def logarApp(user, senha, app):

    keyboard.send('windows')

    time.sleep(0.5)
    pyautogui.write(app)

    time.sleep(1)
    keyboard.send('enter')

    time.sleep(2)
    keyboard.send('tab')
    pyautogui.write(user)

    time.sleep(0.5)
    keyboard.send('tab')
    pyautogui.write(senha)

    time.sleep(0.5)
    keyboard.send('tab')
    keyboard.send('enter')


def salvarProdutos(produto, qtd, armazen, doc, nº):
    time.sleep(1)
    if nº == 0:
        clicar(3, 'tab')
    else:
        clicar(1, 'tab')
    pyautogui.write(produto)

    clicar(1, 'tab')
    pyautogui.write(qtd)

    clicar(5, 'tab')
    pyautogui.write(armazen)

    clicar(12, 'tab')
    pyautogui.write(doc)

    clicar(8, 'tab')
    clicar(1, 'enter')

appName = "\"C:\\Program Files\\Google\\Chrome\\Application\\chrome_proxy.exe\" --profile-directory=\"Profile 1\" --app-id=ginkifbklpdfedpooabndoadkohlhndl"

logarApp(
    "admin",
    "sccp@1910",
    appName
)

for i in range(len(produtolist)):
    salvarProdutos(str(int(produtolist[i])), str(int(qtdlist[i])), str(armazenlist[i]), str(documento), i)
