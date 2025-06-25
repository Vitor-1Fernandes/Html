from openpyxl import load_workbook



planilha = load_workbook('credenciais.xlsx')
pagina = planilha['Plan1']

produto = []
quantidade = []
armazen = []
documento = 123456

for dados in pagina.iter_rows(values_only=True):
    produto.append(dados[0])
    quantidade.append(dados[1])
    armazen.append(dados[2])

